# app.py
from __future__ import annotations

import os
import sqlite3
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo
from typing import Optional, Tuple, Dict, List

from flask import (
    Flask, request, redirect, url_for, session,
    render_template, flash, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook


# ==========================================================
# CONFIG
# ==========================================================
APP_PORT = 5055
DB_FILE = "produccion.db"

TZ = ZoneInfo("America/Tijuana")

DEFAULT_START_TIME = "07:00"
DEFAULT_END_TIME = "17:00"
DEFAULT_GRACE_MINUTES = 0

LINES = ["FORD", "GM", "CHRYSLER", "AUTOPRUEBA", "OTRA"]


def create_app() -> Flask:
    app = Flask(__name__)
    app.secret_key = os.environ.get("APP_SECRET_KEY", "CAMBIA_ESTA_LLAVE_SECRETA_123")

    # Permite usar now() dentro de Jinja templates
    @app.context_processor
    def inject_now():
        return {"now": lambda: datetime.now(TZ)}

    # ------------------------------------------------------
    # DB INIT
    # ------------------------------------------------------
    db_init()
    init_default_settings()
    seed_default_admin()

    # ======================================================
    # ROUTES
    # ======================================================
    @app.route("/")
    def home():
        user = current_user()
        if not user:
            return redirect(url_for("login"))

        role = user["role"]
        if role == "technician":
            return redirect(url_for("technician"))
        if role == "supervisor":
            return redirect(url_for("supervisor"))
        return redirect(url_for("admin"))

    # -------------------------
    # LOGIN / LOGOUT
    # -------------------------
    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            employee_no = (request.form.get("employee_no") or "").strip()
            pin = (request.form.get("pin") or "").strip()

            user = get_user_by_employee(employee_no)
            if not user or user["is_active"] != 1:
                flash("Usuario no encontrado o desactivado.", "err")
                return redirect(url_for("login"))

            if not check_password_hash(user["pin_hash"], pin):
                flash("PIN incorrecto.", "err")
                return redirect(url_for("login"))

            session["employee_no"] = employee_no
            return redirect(url_for("home"))

        return render_template("login.html", user=None)

    @app.route("/logout")
    def logout():
        session.clear()
        return redirect(url_for("login"))

    # -------------------------
    # TECHNICIAN
    # -------------------------
    @app.route("/technician", methods=["GET", "POST"])
    def technician():
        user = require_login_role("technician")
        if not user:
            return redirect(url_for("login"))

        emp = user["employee_no"]
        selected_date = request.args.get("date") or date.today().isoformat()

        schedule = get_capture_window()
        start_str, end_str = schedule.start_str, schedule.end_str
        hours = list(range(schedule.start_h, schedule.end_h + 1))

        now_dt = datetime.now(TZ)
        now_hour = now_dt.hour
        now_time_hhmm = now_dt.strftime("%H:%M")

        # ====== IMPORTANTE ======
        # Quitamos el cierre de día por completo (para evitar bloqueos por error)
        # Si antes se guardó algo en day_closures, ya NO afectará al técnico.
        closed = False
        closed_at = None

        auto_locked = is_auto_locked_for_technician(selected_date)
        can_edit = (not auto_locked)

        if request.method == "POST":
            action = (request.form.get("action") or "add_entry").strip()

            # Si alguien intenta mandar "close_day" por POST, lo ignoramos.
            if action == "close_day":
                flash("Acción deshabilitada: ya no existe el cierre de día.", "err")
                return redirect(url_for("technician", date=selected_date))

            # --------- bloqueo ----------
            # YA NO usamos is_day_closed() aquí.
            if action in {"add_entry", "edit_entry", "delete_entry", "set_inventory"}:
                if is_auto_locked_for_technician(selected_date):
                    flash("Edición bloqueada: fuera del horario permitido.", "err")
                    return redirect(url_for("technician", date=selected_date))

            # --------- inventario inicial ----------
            if action == "set_inventory":
                initial_qty = parse_nonneg_int(request.form.get("initial_qty"))
                if initial_qty is None:
                    flash("Inventario inicial inválido. Debe ser un número 0 o mayor.", "err")
                    return redirect(url_for("technician", date=selected_date))

                upsert_inventory_initial(emp, selected_date, initial_qty)
                flash(f"Inventario inicial guardado: {initial_qty}", "ok")
                return redirect(url_for("technician", date=selected_date))

            # --------- agregar registro ----------
            if action == "add_entry":
                part_no = (request.form.get("part_no") or "").strip().upper()
                qty = parse_pos_int(request.form.get("qty"))
                time_hhmm = (request.form.get("time_hhmm") or "").strip() or now_time_hhmm

                #  Tipo (GOOD/FIX/SCRAP)
                log_type = normalize_log_type(request.form.get("log_type"))

                if not parse_time_hhmm(time_hhmm):
                    flash("Hora inválida. Usa formato HH:MM.", "err")
                    return redirect(url_for("technician", date=selected_date))

                if not is_time_within_window(time_hhmm, start_str, end_str):
                    flash(f"Solo se permite registrar de {start_str} a {end_str}.", "err")
                    return redirect(url_for("technician", date=selected_date))

                if not part_no:
                    flash("Número de Parte es obligatorio.", "err")
                    return redirect(url_for("technician", date=selected_date))

                if qty is None:
                    flash("Cantidad debe ser un número mayor a 0.", "err")
                    return redirect(url_for("technician", date=selected_date))

                hour = int(time_hhmm.split(":")[0])
                insert_part_log(emp, selected_date, time_hhmm, hour, part_no, qty, log_type)

                flash(f"Registro guardado: {time_hhmm} | {part_no} | {log_type} | qty {qty}", "ok")
                return redirect(url_for("technician", date=selected_date))

            # --------- editar registro ----------
            if action == "edit_entry":
                entry_id = parse_pos_int(request.form.get("entry_id"))
                part_no = (request.form.get("part_no") or "").strip().upper()
                qty = parse_pos_int(request.form.get("qty"))
                time_hhmm = (request.form.get("time_hhmm") or "").strip()

                #  Tipo (GOOD/FIX/SCRAP)
                log_type = normalize_log_type(request.form.get("log_type"))

                if entry_id is None:
                    flash("Registro inválido.", "err")
                    return redirect(url_for("technician", date=selected_date))

                if not parse_time_hhmm(time_hhmm):
                    flash("Hora inválida. Usa formato HH:MM.", "err")
                    return redirect(url_for("technician", date=selected_date))

                if not is_time_within_window(time_hhmm, start_str, end_str):
                    flash(f"Solo se permite registrar de {start_str} a {end_str}.", "err")
                    return redirect(url_for("technician", date=selected_date))

                if not part_no:
                    flash("Número de Parte es obligatorio.", "err")
                    return redirect(url_for("technician", date=selected_date))

                if qty is None:
                    flash("Cantidad debe ser un número mayor a 0.", "err")
                    return redirect(url_for("technician", date=selected_date))

                hour = int(time_hhmm.split(":")[0])
                update_part_log(entry_id, emp, selected_date, time_hhmm, hour, part_no, qty, log_type)

                flash("Registro actualizado correctamente.", "ok")
                return redirect(url_for("technician", date=selected_date))

            # --------- borrar registro ----------
            if action == "delete_entry":
                entry_id = parse_pos_int(request.form.get("entry_id"))
                if entry_id is None:
                    flash("Registro inválido.", "err")
                    return redirect(url_for("technician", date=selected_date))

                delete_part_log(entry_id, emp, selected_date)
                flash("Registro eliminado correctamente.", "ok")
                return redirect(url_for("technician", date=selected_date))

        # --------- GET: data ----------
        entries = get_part_logs_by_emp_date(emp, selected_date, start_str, end_str)

        # Resumen por hora (incluye TODO: GOOD+FIX+SCRAP)
        hour_summary = {h: 0 for h in hours}
        for e in entries:
            hh = int(e["hour"])
            if hh in hour_summary:
                hour_summary[hh] += int(e["qty"])

        # Metas (GOOD+FIX)
        tech_goal = get_tech_goal(emp)
        tech_total = calc_tech_total(emp, selected_date, start_str, end_str)
        tech_remaining = None if tech_goal is None else max(0, tech_goal - tech_total)

        # Inventario (TODO descuenta)
        inv_initial = get_inventory_initial(emp, selected_date)
        inv_produced = calc_inventory_produced(emp, selected_date, start_str, end_str)
        inv_final = None if inv_initial is None else (inv_initial - inv_produced)

        return render_template(
            "technician.html",
            user=user,
            selected_date=selected_date,
            hours=hours,
            now_hour=now_hour,
            now_time_hhmm=now_time_hhmm,
            can_edit=can_edit,
            is_closed=closed,
            closed_at=closed_at,
            auto_locked=auto_locked,
            schedule_start=start_str,
            schedule_end=end_str,
            schedule_grace=schedule.grace_minutes,
            entries=entries,
            hour_summary=hour_summary,
            tech_goal=tech_goal,
            tech_total=tech_total,
            tech_remaining=tech_remaining,
            inv_initial=inv_initial,
            inv_produced=inv_produced,
            inv_final=inv_final,
        )

    # -------------------------
    # SUPERVISOR
    # -------------------------
    @app.route("/supervisor", methods=["GET", "POST"], endpoint="supervisor")
    def supervisor_panel():
        user = require_login_role("supervisor", "admin")
        if not user:
            return redirect(url_for("login"))

        schedule = get_capture_window()
        start_str, end_str = schedule.start_str, schedule.end_str

        if request.method == "POST":
            action = (request.form.get("action") or "").strip()

            if action == "set_tech_goal":
                emp_no = (request.form.get("employee_no") or "").strip()
                goal = parse_nonneg_int(request.form.get("daily_goal"))

                if not emp_no or goal is None:
                    flash("Meta de técnico inválida.", "err")
                    return redirect(url_for("supervisor"))

                upsert_tech_goal(emp_no, goal)
                flash(f"Meta asignada al técnico {emp_no}: {goal}", "ok")
                return redirect(url_for("supervisor"))

            if action == "set_line_goal":
                line = (request.form.get("line") or "").strip().upper()
                goal = parse_nonneg_int(request.form.get("daily_goal"))

                if not line or goal is None:
                    flash("Meta de línea inválida.", "err")
                    return redirect(url_for("supervisor"))

                upsert_line_goal(line, goal)
                flash(f"Meta asignada a la línea {line}: {goal}", "ok")
                return redirect(url_for("supervisor"))

        selected_date = request.args.get("date") or date.today().isoformat()
        selected_emp = (request.args.get("emp") or "").strip()
        selected_line = (request.args.get("line") or "").strip().upper()

        users_list = get_active_users()

        tech_goals_map = get_tech_goals_map()
        line_goals_map = get_line_goals_map()

        #  Bitácora visible (incluye log_type para GOOD/FIX/SCRAP)
        logs = get_supervisor_logs(selected_date, start_str, end_str, selected_emp, selected_line)

        # Resumen por técnico (meta: GOOD+FIX)
        summary_by_emp = get_summary_by_emp(selected_date, start_str, end_str)

        # Resumen por línea (meta: GOOD+FIX)
        summary_by_line = get_summary_by_line(selected_date, start_str, end_str)

        # Inventario por línea (descunta TODO)
        inv_lines, inv_by_line = get_inventory_line_view(selected_date, start_str, end_str)

        # Progresos de metas
        emp_progress = []
        for r in summary_by_emp:
            emp_no = r["employee_no"]
            total = int(r["total_qty"] or 0)
            goal = tech_goals_map.get(emp_no)
            remaining = None if goal is None else max(0, goal - total)
            emp_progress.append({
                "employee_no": emp_no,
                "name": r["name"],
                "line": r["line"],
                "total": total,
                "goal": goal,
                "remaining": remaining
            })

        line_progress = []
        for r in summary_by_line:
            line = (r["line"] or "").upper() or "SIN ASIGNAR"
            total = int(r["total_qty"] or 0)
            goal = line_goals_map.get(line)
            remaining = None if goal is None else max(0, goal - total)
            line_progress.append({
                "line": line,
                "total": total,
                "goal": goal,
                "remaining": remaining
            })

        return render_template(
            "supervisor.html",
            user=user,
            selected_date=selected_date,
            selected_emp=selected_emp,
            selected_line=selected_line,
            users_list=users_list,
            logs=logs,
            lines=LINES,
            emp_progress=emp_progress,
            line_progress=line_progress,
            schedule_start=start_str,
            schedule_end=end_str,
            inv_lines=inv_lines,
            inv_by_line=inv_by_line,
        )

    # -------------------------
    # DASHBOARD POR LÍNEA
    # -------------------------
    @app.route("/line_dashboard", methods=["GET"])
    def line_dashboard():
        user = require_login_role("supervisor", "admin")
        if not user:
            return redirect(url_for("login"))

        schedule = get_capture_window()
        start_str, end_str = schedule.start_str, schedule.end_str

        selected_date = request.args.get("date") or date.today().isoformat()

        totals_by_line = get_summary_by_line(selected_date, start_str, end_str)
        line_goals_map = get_line_goals_map()

        tech_by_line = get_tech_by_line(selected_date, start_str, end_str)

        tot_map = {(r["line"] or "SIN ASIGNAR").upper(): int(r["total_qty"] or 0) for r in totals_by_line}
        for ln in LINES:
            ln2 = (ln or "").upper()
            tot_map.setdefault(ln2, 0)

        line_cards = []
        for line, total in tot_map.items():
            goal = line_goals_map.get(line)
            remaining = None if goal is None else max(0, goal - total)

            pct = None
            if goal is not None and goal > 0:
                pct = min(100, round((total * 100) / goal))

            line_cards.append({
                "line": line,
                "total": total,
                "goal": goal,
                "remaining": remaining,
                "pct": pct
            })

        line_cards.sort(key=lambda x: x["total"], reverse=True)

        tech_map: Dict[str, List[dict]] = {}
        for t in tech_by_line:
            line = (t["line"] or "SIN ASIGNAR").upper()
            tech_map.setdefault(line, []).append({
                "employee_no": t["employee_no"],
                "name": t["name"],
                "total": int(t["total_qty"] or 0)
            })

        return render_template(
            "line_dashboard.html",
            user=user,
            selected_date=selected_date,
            schedule_start=start_str,
            schedule_end=end_str,
            line_cards=line_cards,
            tech_map=tech_map,
        )

    # -------------------------
    # REPORTE SEMANAL (Lunes a Viernes)
    # -------------------------
    @app.route("/weekly_report", methods=["GET"])
    def weekly_report():
        user = require_login_role("supervisor", "admin")
        if not user:
            return redirect(url_for("login"))

        schedule = get_capture_window()
        start_str, end_str = schedule.start_str, schedule.end_str

        # Ahora week_end es VIERNES
        year, week, week_start, week_end = parse_year_week_request()
        workdays_per_week = 5

        techs = get_active_techs()

        totals_emp_map = get_week_totals_emp_map(week_start, week_end, start_str, end_str)  # GOOD+FIX
        tech_goal_map = get_tech_goals_map()

        totals_line_map = get_week_totals_line_map(week_start, week_end, start_str, end_str)  # GOOD+FIX
        line_goal_map = get_line_goals_map()

        tech_rows = []
        for t in techs:
            emp = t["employee_no"]
            name = t["name"]
            line = (t["line"] or "SIN ASIGNAR").upper()

            total = totals_emp_map.get(emp, 0)

            daily_goal = tech_goal_map.get(emp)
            weekly_goal = None if daily_goal is None else daily_goal * workdays_per_week
            remaining = None if weekly_goal is None else max(0, weekly_goal - total)

            eff = None
            if weekly_goal is not None and weekly_goal > 0:
                eff = round((total * 100) / weekly_goal)

            tech_rows.append({
                "employee_no": emp,
                "name": name,
                "line": line,
                "total": total,
                "weekly_goal": weekly_goal,
                "remaining": remaining,
                "eff": eff,
            })

        line_rows = []
        for ln in LINES:
            ln2 = (ln or "").upper()
            total = totals_line_map.get(ln2, 0)

            daily_goal = line_goal_map.get(ln2)
            weekly_goal = None if daily_goal is None else daily_goal * workdays_per_week
            remaining = None if weekly_goal is None else max(0, weekly_goal - total)

            eff = None
            if weekly_goal is not None and weekly_goal > 0:
                eff = round((total * 100) / weekly_goal)

            line_rows.append({
                "line": ln2,
                "total": total,
                "weekly_goal": weekly_goal,
                "remaining": remaining,
                "eff": eff
            })

        tech_rows.sort(key=lambda x: (x["eff"] if x["eff"] is not None else -1, x["total"]), reverse=True)

        # Comparativo semana anterior (Lunes a Viernes)
        prev_start = week_start - timedelta(days=7)
        prev_end = prev_start + timedelta(days=4)  # Lunes–Viernes

        prev_emp_map = get_week_totals_emp_map(prev_start, prev_end, start_str, end_str)
        prev_line_map = get_week_totals_line_map(prev_start, prev_end, start_str, end_str)

        current_total_all = sum([r["total"] for r in tech_rows])
        prev_total_all = sum(prev_emp_map.values())
        delta_total_all = current_total_all - prev_total_all

        compare_lines = []
        for ln in LINES:
            ln2 = (ln or "").upper()
            curr = totals_line_map.get(ln2, 0)
            prevv = prev_line_map.get(ln2, 0)
            compare_lines.append({
                "line": ln2,
                "current": curr,
                "previous": prevv,
                "delta": curr - prevv
            })
        compare_lines.sort(key=lambda x: x["current"], reverse=True)

        return render_template(
            "weekly_report.html",
            user=user,
            year=year,
            week=week,
            week_start=week_start,
            week_end=week_end,
            schedule_start=start_str,
            schedule_end=end_str,
            workdays_per_week=workdays_per_week,
            tech_rows=tech_rows,
            line_rows=line_rows,
            prev_start=prev_start,
            prev_end=prev_end,
            prev_total_all=prev_total_all,
            current_total_all=current_total_all,
            delta_total_all=delta_total_all,
            compare_lines=compare_lines,
        )

    # -------------------------
    # EXPORT EXCEL REPORTE SEMANAL (Lunes a Viernes)
    # -------------------------
    @app.route("/export_weekly_excel", methods=["GET"])
    def export_weekly_excel():
        user = require_login_role("supervisor", "admin")
        if not user:
            return redirect(url_for("login"))

        schedule = get_capture_window()
        start_str, end_str = schedule.start_str, schedule.end_str

        # Ahora week_end es VIERNES
        year, week, week_start, week_end = parse_year_week_request()
        workdays_per_week = 5

        techs = get_active_techs()

        totals_emp_map = get_week_totals_emp_map(week_start, week_end, start_str, end_str)  # GOOD+FIX
        tech_goal_map = get_tech_goals_map()

        totals_line_map = get_week_totals_line_map(week_start, week_end, start_str, end_str)  # GOOD+FIX
        line_goal_map = get_line_goals_map()

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Tecnicos"
        ws1.append(["Empleado", "Nombre", "Linea", "Total Semana (GOOD+FIX)", "Meta Semana", "Faltante", "Eficiencia"])

        for t in techs:
            emp = t["employee_no"]
            name = t["name"]
            line = (t["line"] or "SIN ASIGNAR").upper()
            total = totals_emp_map.get(emp, 0)

            daily_goal = tech_goal_map.get(emp)
            weekly_goal = None if daily_goal is None else daily_goal * workdays_per_week
            remaining = None if weekly_goal is None else max(0, weekly_goal - total)

            eff = None
            if weekly_goal is not None and weekly_goal > 0:
                eff = round((total * 100) / weekly_goal)

            ws1.append([
                emp, name, line, total,
                weekly_goal if weekly_goal is not None else "",
                remaining if remaining is not None else "",
                f"{eff}%" if eff is not None else ""
            ])

        ws2 = wb.create_sheet("Lineas")
        ws2.append(["Linea", "Total Semana (GOOD+FIX)", "Meta Semana", "Faltante", "Eficiencia"])

        for ln in LINES:
            ln2 = (ln or "").upper()
            total = totals_line_map.get(ln2, 0)

            daily_goal = line_goal_map.get(ln2)
            weekly_goal = None if daily_goal is None else daily_goal * workdays_per_week
            remaining = None if weekly_goal is None else max(0, weekly_goal - total)

            eff = None
            if weekly_goal is not None and weekly_goal > 0:
                eff = round((total * 100) / weekly_goal)

            ws2.append([
                ln2, total,
                weekly_goal if weekly_goal is not None else "",
                remaining if remaining is not None else "",
                f"{eff}%" if eff is not None else ""
            ])

        autosize_worksheets(wb)

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"Eficiencia_Semanal_{year}_W{week}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # -------------------------
    # EXPORT EXCEL DIARIO (SEPARA GOOD/FIX/SCRAP)
    # -------------------------
    @app.route("/export_excel", methods=["GET"])
    def export_excel():
        user = require_login_role("supervisor", "admin")
        if not user:
            return redirect(url_for("login"))

        schedule = get_capture_window()
        start_str, end_str = schedule.start_str, schedule.end_str

        selected_date = request.args.get("date") or date.today().isoformat()
        selected_emp = (request.args.get("emp") or "").strip()
        selected_line = (request.args.get("line") or "").strip().upper()

        rows = get_export_daily_rows(selected_date, start_str, end_str, selected_emp, selected_line)

        wb = Workbook()
        ws = wb.active
        ws.title = "Bitacora"

        ws.append(["Fecha", "Hora", "Empleado", "Nombre", "Línea", "Numero de Parte", "Tipo", "GOOD", "FIX", "SCRAP", "Actualización"])

        for r in rows:
            t = normalize_log_type(r["log_type"])
            good = r["qty"] if t == "GOOD" else ""
            fix = r["qty"] if t == "FIX" else ""
            scrap = r["qty"] if t == "SCRAP" else ""

            ws.append([
                r["log_date"],
                r["log_time"],
                r["employee_no"],
                r["name"],
                r["line"],
                r["part_no"],
                t,
                good,
                fix,
                scrap,
                r["updated_at"],
            ])

        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            ws.column_dimensions[col].width = 18

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"Bitacora_{selected_date}.xlsx"
        return send_file(
            bio,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # -------------------------
    # ADMIN
    # -------------------------
    @app.route("/admin", methods=["GET", "POST"])
    def admin():
        user = require_login_role("admin")
        if not user:
            return redirect(url_for("login"))

        if request.method == "POST":
            action = (request.form.get("action") or "").strip()

            # Actualizar horario
            if action == "update_schedule":
                start_time = (request.form.get("start_time") or "").strip()
                end_time = (request.form.get("end_time") or "").strip()
                grace_raw = (request.form.get("grace_minutes") or "").strip()

                if not parse_time_hhmm(start_time) or not parse_time_hhmm(end_time):
                    flash("Horario inválido. Usa HH:MM.", "err")
                    return redirect(url_for("admin"))

                if start_time > end_time:
                    flash("La hora de inicio no puede ser mayor que la hora final.", "err")
                    return redirect(url_for("admin"))

                grace = parse_nonneg_int(grace_raw)
                if grace is None:
                    grace = 0
                grace = min(300, grace)

                set_setting("start_time", start_time)
                set_setting("end_time", end_time)
                set_setting("grace_minutes", str(grace))

                flash(f"Horario actualizado: {start_time} a {end_time} (gracia {grace} min).", "ok")
                return redirect(url_for("admin"))

            # Acciones que requieren DB abierta
            with db() as conn:
                cur = conn.cursor()

                # Crear usuario
                if action == "create_user":
                    employee_no = (request.form.get("employee_no") or "").strip()
                    name = (request.form.get("name") or "").strip()
                    role = (request.form.get("role") or "technician").strip()
                    pin = (request.form.get("pin") or "").strip()
                    line = (request.form.get("line") or "").strip().upper()

                    if line == "OTRA":
                        line = (request.form.get("line_other") or "").strip().upper()

                    if not employee_no or not name or not pin:
                        flash("Empleado, nombre y PIN son obligatorios.", "err")
                        return redirect(url_for("admin"))
                    if role not in ("technician", "supervisor", "admin"):
                        flash("Rol inválido.", "err")
                        return redirect(url_for("admin"))

                    try:
                        now = datetime.now(TZ).isoformat(timespec="seconds")
                        cur.execute("""
                            INSERT INTO users (employee_no, name, role, pin_hash, is_active, created_at, line)
                            VALUES (?, ?, ?, ?, 1, ?, ?)
                        """, (employee_no, name, role, generate_password_hash(pin), now, line))
                        conn.commit()
                        flash("Usuario creado correctamente.", "ok")
                    except sqlite3.IntegrityError:
                        flash("Ese número de empleado ya existe.", "err")

                    return redirect(url_for("admin"))

                # Activar / desactivar
                if action == "toggle_active":
                    employee_no = (request.form.get("employee_no") or "").strip()
                    if employee_no == "0000":
                        flash("No se puede desactivar el admin principal.", "err")
                        return redirect(url_for("admin"))

                    cur.execute("""
                        UPDATE users
                        SET is_active = CASE WHEN is_active=1 THEN 0 ELSE 1 END
                        WHERE employee_no = ?
                    """, (employee_no,))
                    conn.commit()
                    flash("Estado actualizado.", "ok")
                    return redirect(url_for("admin"))

                # Eliminar usuario seguro (si tiene historial => desactiva)
                if action == "delete_user":
                    employee_no = (request.form.get("employee_no") or "").strip()
                    res = safe_delete_user(cur, employee_no)
                    conn.commit()
                    flash(res["msg"], res["cat"])
                    return redirect(url_for("admin"))

                # Eliminar full (borra historial)
                if action == "delete_user_full":
                    employee_no = (request.form.get("employee_no") or "").strip()
                    res = full_delete_user(cur, employee_no)
                    conn.commit()
                    flash(res["msg"], res["cat"])
                    return redirect(url_for("admin"))

                # Reset PIN
                if action == "reset_pin":
                    employee_no = (request.form.get("employee_no") or "").strip()
                    new_pin = (request.form.get("new_pin") or "").strip()
                    if not employee_no or not new_pin:
                        flash("Número de empleado y nuevo PIN son obligatorios.", "err")
                        return redirect(url_for("admin"))

                    cur.execute(
                        "UPDATE users SET pin_hash = ? WHERE employee_no = ?",
                        (generate_password_hash(new_pin), employee_no)
                    )
                    conn.commit()
                    flash("PIN actualizado.", "ok")
                    return redirect(url_for("admin"))

                # Cambiar línea
                if action == "update_line":
                    employee_no = (request.form.get("employee_no") or "").strip()
                    new_line = (request.form.get("new_line") or "").strip().upper()

                    if new_line == "OTRA":
                        new_line = (request.form.get("new_line_other") or "").strip().upper()

                    cur.execute(
                        "UPDATE users SET line = ? WHERE employee_no = ?",
                        (new_line, employee_no)
                    )
                    conn.commit()
                    flash("Línea actualizada.", "ok")
                    return redirect(url_for("admin"))

                flash("Acción no reconocida.", "err")
                return redirect(url_for("admin"))

        users_list = get_all_users()
        schedule = get_capture_window()

        return render_template(
            "admin.html",
            user=user,
            users_list=users_list,
            lines=LINES,
            schedule_start=schedule.start_str,
            schedule_end=schedule.end_str,
            schedule_grace=schedule.grace_minutes,
        )

    # -------------------------
    # ADMIN LOGS (BORRAR CAPTURAS AUNQUE ESTÉ CERRADO)
    # -------------------------
    @app.route("/admin_logs", methods=["GET", "POST"])
    def admin_logs():
        user = require_login_role("admin")
        if not user:
            return redirect(url_for("login"))

        schedule = get_capture_window()
        start_str, end_str = schedule.start_str, schedule.end_str

        if request.method == "POST":
            action = (request.form.get("action") or "").strip()
            if action == "delete_log":
                entry_id = parse_pos_int(request.form.get("entry_id"))
                if entry_id is None:
                    flash("ID inválido.", "err")
                    return redirect(url_for("admin_logs"))

                row = get_part_log_by_id(entry_id)
                if not row:
                    flash("Ese registro no existe.", "err")
                    return redirect(url_for("admin_logs"))

                delete_part_log_any(entry_id)
                flash(
                    f"Captura eliminada: ID {row['id']} | {row['employee_no']} | {row['log_date']} {row['log_time']} | {row['part_no']} | {normalize_log_type(row['log_type'])} | qty {row['qty']}",
                    "ok"
                )
                return redirect(url_for("admin_logs", date=row["log_date"], emp=row["employee_no"]))

            flash("Acción no reconocida.", "err")
            return redirect(url_for("admin_logs"))

        selected_date = request.args.get("date") or date.today().isoformat()
        selected_emp = (request.args.get("emp") or "").strip()
        selected_line = (request.args.get("line") or "").strip().upper()

        users_list = get_active_users()
        logs = get_admin_logs(selected_date, selected_emp, selected_line)

        return render_template(
            "admin_logs.html",
            user=user,
            selected_date=selected_date,
            selected_emp=selected_emp,
            selected_line=selected_line,
            users_list=users_list,
            lines=LINES,
            logs=logs,
            schedule_start=start_str,
            schedule_end=end_str,
        )

    return app


# ==========================================================
# DB CORE
# ==========================================================
def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


class db:
    """Context manager simple para abrir/cerrar DB sin repetir."""
    def __enter__(self):
        self.conn = db_connect()
        return self.conn

    def __exit__(self, exc_type, exc, tb):
        self.conn.close()


def column_exists(conn: sqlite3.Connection, table: str, col: str) -> bool:
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = [r["name"] for r in cur.fetchall()]
    return col in cols


def db_init() -> None:
    with db() as conn:
        cur = conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_no TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                role TEXT NOT NULL CHECK(role IN ('technician','supervisor','admin')),
                pin_hash TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                line TEXT NOT NULL DEFAULT ''
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)

        # Incluye log_type desde el CREATE TABLE (para DBs nuevas)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS part_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_no TEXT NOT NULL,
                log_date TEXT NOT NULL,
                log_time TEXT NOT NULL,
                hour INTEGER NOT NULL,
                part_no TEXT NOT NULL,
                qty INTEGER NOT NULL,
                log_type TEXT NOT NULL DEFAULT 'GOOD',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(employee_no) REFERENCES users(employee_no)
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS day_closures (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_no TEXT NOT NULL,
                log_date TEXT NOT NULL,
                is_closed INTEGER NOT NULL DEFAULT 0,
                closed_at TEXT,
                UNIQUE(employee_no, log_date),
                FOREIGN KEY(employee_no) REFERENCES users(employee_no)
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS tech_goals (
                employee_no TEXT PRIMARY KEY,
                daily_goal INTEGER NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(employee_no) REFERENCES users(employee_no)
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS line_goals (
                line TEXT PRIMARY KEY,
                daily_goal INTEGER NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS tech_inventory (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_no TEXT NOT NULL,
                log_date TEXT NOT NULL,
                initial_qty INTEGER NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(employee_no, log_date),
                FOREIGN KEY(employee_no) REFERENCES users(employee_no)
            )
        """)

        conn.commit()

        # MIGRACIONES (por si existían DBs viejas)
        if not column_exists(conn, "users", "line"):
            cur.execute("ALTER TABLE users ADD COLUMN line TEXT NOT NULL DEFAULT ''")
            conn.commit()

        if not column_exists(conn, "part_logs", "log_type"):
            cur.execute("ALTER TABLE part_logs ADD COLUMN log_type TEXT NOT NULL DEFAULT 'GOOD'")
            conn.commit()
            cur.execute("UPDATE part_logs SET log_type='GOOD' WHERE log_type IS NULL OR log_type=''")
            conn.commit()


def get_setting(key: str, default: str) -> str:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT value FROM app_settings WHERE key = ?", (key,))
        row = cur.fetchone()
        return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO app_settings (key, value)
            VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value
        """, (key, value))
        conn.commit()


def init_default_settings() -> None:
    if get_setting("start_time", "") == "":
        set_setting("start_time", DEFAULT_START_TIME)
    if get_setting("end_time", "") == "":
        set_setting("end_time", DEFAULT_END_TIME)
    if get_setting("grace_minutes", "") == "":
        set_setting("grace_minutes", str(DEFAULT_GRACE_MINUTES))


def seed_default_admin() -> None:
    """
    Usuarios por defecto si la BD está vacía:
    - Admin: 0000 / 0000
    - Supervisor: 9999 / 9999
    """
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) AS c FROM users")
        c = int(cur.fetchone()["c"] or 0)

        if c > 0:
            return

        now = datetime.now(TZ).isoformat(timespec="seconds")

        cur.execute("""
            INSERT INTO users (employee_no, name, role, pin_hash, is_active, created_at, line)
            VALUES (?, ?, ?, ?, 1, ?, ?)
        """, ("0000", "ADMIN", "admin", generate_password_hash("0000"), now, ""))

        cur.execute("""
            INSERT INTO users (employee_no, name, role, pin_hash, is_active, created_at, line)
            VALUES (?, ?, ?, ?, 1, ?, ?)
        """, ("9999", "SUPERVISOR", "supervisor", generate_password_hash("9999"), now, ""))

        conn.commit()


# ==========================================================
# TIME / SCHEDULE
# ==========================================================
def parse_time_hhmm(t: str) -> Optional[Tuple[int, int]]:
    try:
        t = (t or "").strip()
        hh, mm = t.split(":")
        h = int(hh)
        m = int(mm)
        if 0 <= h <= 23 and 0 <= m <= 59:
            return h, m
    except Exception:
        return None
    return None


@dataclass(frozen=True)
class CaptureWindow:
    start_str: str
    end_str: str
    grace_minutes: int
    start_h: int
    start_m: int
    end_h: int
    end_m: int


def get_capture_window() -> CaptureWindow:
    start_str = get_setting("start_time", DEFAULT_START_TIME)
    end_str = get_setting("end_time", DEFAULT_END_TIME)
    grace_str = get_setting("grace_minutes", str(DEFAULT_GRACE_MINUTES))

    st = parse_time_hhmm(start_str) or parse_time_hhmm(DEFAULT_START_TIME)
    et = parse_time_hhmm(end_str) or parse_time_hhmm(DEFAULT_END_TIME)

    grace = parse_nonneg_int(grace_str)
    if grace is None:
        grace = DEFAULT_GRACE_MINUTES
    grace = min(300, grace)

    return CaptureWindow(
        start_str=start_str,
        end_str=end_str,
        grace_minutes=grace,
        start_h=st[0],
        start_m=st[1],
        end_h=et[0],
        end_m=et[1],
    )


def is_time_within_window(log_time_hhmm: str, start_str: str, end_str: str) -> bool:
    return bool(log_time_hhmm) and (start_str <= log_time_hhmm <= end_str)


# ==========================================================
# AUTH
# ==========================================================
def get_user_by_employee(employee_no: str):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE employee_no = ?", (employee_no,))
        return cur.fetchone()


def current_user():
    emp = session.get("employee_no")
    if not emp:
        return None
    return get_user_by_employee(emp)


def require_login_role(*roles):
    user = current_user()
    if not user or user["is_active"] != 1:
        return None
    if roles and user["role"] not in roles:
        flash("No tienes permisos para esa sección.", "err")
        return None
    return user


# ==========================================================
# LOCK RULES
# ==========================================================
def is_day_closed(employee_no: str, log_date: str) -> bool:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT is_closed FROM day_closures
            WHERE employee_no = ? AND log_date = ?
        """, (employee_no, log_date))
        row = cur.fetchone()
        return bool(row is not None and row["is_closed"] == 1)


def get_day_closure(employee_no: str, log_date: str):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT is_closed, closed_at
            FROM day_closures
            WHERE employee_no = ? AND log_date = ?
        """, (employee_no, log_date))
        return cur.fetchone()


def close_day(employee_no: str, log_date: str) -> None:
    now = datetime.now(TZ).isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO day_closures (employee_no, log_date, is_closed, closed_at)
                VALUES (?, ?, 1, ?)
            """, (employee_no, log_date, now))
        except sqlite3.IntegrityError:
            cur.execute("""
                UPDATE day_closures
                SET is_closed = 1, closed_at = ?
                WHERE employee_no = ? AND log_date = ?
            """, (now, employee_no, log_date))
        conn.commit()


def is_auto_locked_for_technician(log_date: str) -> bool:
    today = date.today().isoformat()
    now = datetime.now(TZ)

    schedule = get_capture_window()

    if log_date < today:
        return True

    if log_date == today:
        end_dt = now.replace(
            hour=schedule.end_h, minute=schedule.end_m, second=59, microsecond=0
        ) + timedelta(minutes=schedule.grace_minutes)
        if now > end_dt:
            return True

    return False


# ==========================================================
# NORMALIZERS / PARSERS
# ==========================================================
def normalize_log_type(value: Optional[str]) -> str:
    t = (value or "GOOD").strip().upper()
    if t not in ("GOOD", "FIX", "SCRAP"):
        t = "GOOD"
    return t


def parse_pos_int(x) -> Optional[int]:
    try:
        v = int(str(x).strip())
        return v if v > 0 else None
    except Exception:
        return None


def parse_nonneg_int(x) -> Optional[int]:
    try:
        s = str(x).strip()
        if s == "":
            return None
        v = int(s)
        return v if v >= 0 else None
    except Exception:
        return None


# ==========================================================
# GOALS / INVENTORY
# ==========================================================
def get_tech_goal(employee_no: str) -> Optional[int]:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT daily_goal FROM tech_goals WHERE employee_no = ?", (employee_no,))
        row = cur.fetchone()
        return int(row["daily_goal"]) if row else None


def upsert_tech_goal(employee_no: str, daily_goal: int) -> None:
    now = datetime.now(TZ).isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO tech_goals (employee_no, daily_goal, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(employee_no) DO UPDATE
            SET daily_goal=excluded.daily_goal, updated_at=excluded.updated_at
        """, (employee_no, daily_goal, now))
        conn.commit()


def upsert_line_goal(line: str, daily_goal: int) -> None:
    now = datetime.now(TZ).isoformat(timespec="seconds")
    line = (line or "").strip().upper()
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO line_goals (line, daily_goal, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(line) DO UPDATE
            SET daily_goal=excluded.daily_goal, updated_at=excluded.updated_at
        """, (line, daily_goal, now))
        conn.commit()


def get_inventory_initial(employee_no: str, log_date: str) -> Optional[int]:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT initial_qty
            FROM tech_inventory
            WHERE employee_no = ? AND log_date = ?
        """, (employee_no, log_date))
        row = cur.fetchone()
        return int(row["initial_qty"]) if row else None


def upsert_inventory_initial(employee_no: str, log_date: str, initial_qty: int) -> None:
    now = datetime.now(TZ).isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO tech_inventory (employee_no, log_date, initial_qty, updated_at)
                VALUES (?, ?, ?, ?)
            """, (employee_no, log_date, initial_qty, now))
        except sqlite3.IntegrityError:
            cur.execute("""
                UPDATE tech_inventory
                SET initial_qty = ?, updated_at = ?
                WHERE employee_no = ? AND log_date = ?
            """, (initial_qty, now, employee_no, log_date))
        conn.commit()


# ==========================================================
# LOGS CRUD
# ==========================================================
def insert_part_log(employee_no: str, log_date: str, log_time: str, hour: int,
                    part_no: str, qty: int, log_type: str) -> None:
    now = datetime.now(TZ).isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO part_logs (employee_no, log_date, log_time, hour, part_no, qty, log_type, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (employee_no, log_date, log_time, hour, part_no, qty, log_type, now, now))
        conn.commit()


def update_part_log(entry_id: int, employee_no: str, log_date: str, log_time: str, hour: int,
                    part_no: str, qty: int, log_type: str) -> None:
    now = datetime.now(TZ).isoformat(timespec="seconds")
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            UPDATE part_logs
            SET part_no = ?, qty = ?, log_time = ?, hour = ?, log_type = ?, updated_at = ?
            WHERE id = ? AND employee_no = ? AND log_date = ?
        """, (part_no, qty, log_time, hour, log_type, now, entry_id, employee_no, log_date))
        conn.commit()


def delete_part_log(entry_id: int, employee_no: str, log_date: str) -> None:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            DELETE FROM part_logs
            WHERE id = ? AND employee_no = ? AND log_date = ?
        """, (entry_id, employee_no, log_date))
        conn.commit()


def get_part_log_by_id(entry_id: int):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, employee_no, log_date, log_time, part_no, qty, log_type
            FROM part_logs
            WHERE id = ?
        """, (entry_id,))
        return cur.fetchone()


def delete_part_log_any(entry_id: int) -> None:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM part_logs WHERE id = ?", (entry_id,))
        conn.commit()


def get_part_logs_by_emp_date(employee_no: str, log_date: str, start_str: str, end_str: str):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT *
            FROM part_logs
            WHERE employee_no = ? AND log_date = ?
              AND log_time BETWEEN ? AND ?
            ORDER BY log_time ASC, id ASC
        """, (employee_no, log_date, start_str, end_str))
        return cur.fetchall()


# ==========================================================
# CALCULATIONS
# ==========================================================
def calc_tech_total(employee_no: str, log_date: str, start_str: str, end_str: str) -> int:
    """Meta del técnico: SOLO GOOD + FIX"""
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(SUM(qty),0) AS total
            FROM part_logs
            WHERE employee_no = ? AND log_date = ?
              AND log_time BETWEEN ? AND ?
              AND UPPER(COALESCE(log_type,'GOOD')) IN ('GOOD','FIX')
        """, (employee_no, log_date, start_str, end_str))
        return int(cur.fetchone()["total"] or 0)


def calc_inventory_produced(employee_no: str, log_date: str, start_str: str, end_str: str) -> int:
    """Inventario: descuenta TODO (GOOD + FIX + SCRAP)"""
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(SUM(qty),0) AS total
            FROM part_logs
            WHERE employee_no = ? AND log_date = ?
              AND log_time BETWEEN ? AND ?
        """, (employee_no, log_date, start_str, end_str))
        return int(cur.fetchone()["total"] or 0)


# ==========================================================
# USERS LISTS
# ==========================================================
def get_active_users():
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT employee_no, name, role, line
            FROM users
            WHERE is_active=1
            ORDER BY role, employee_no
        """)
        return cur.fetchall()


def get_all_users():
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT employee_no, name, role, line, is_active, created_at
            FROM users
            ORDER BY role, employee_no
        """)
        return cur.fetchall()


def get_active_techs():
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT employee_no, name, line
            FROM users
            WHERE role='technician' AND is_active=1
            ORDER BY UPPER(line), employee_no
        """)
        return cur.fetchall()


def get_tech_goals_map() -> Dict[str, int]:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT employee_no, daily_goal FROM tech_goals")
        return {r["employee_no"]: int(r["daily_goal"]) for r in cur.fetchall()}


def get_line_goals_map() -> Dict[str, int]:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("SELECT line, daily_goal FROM line_goals")
        return {r["line"].upper(): int(r["daily_goal"]) for r in cur.fetchall()}


# ==========================================================
# SUPERVISOR QUERIES
# ==========================================================
def get_supervisor_logs(selected_date: str, start_str: str, end_str: str,
                        selected_emp: str, selected_line: str):
    # incluir pl.log_type para reflejar GOOD/FIX/SCRAP en supervisor.html
    query = """
        SELECT pl.id, pl.log_date, pl.log_time, pl.employee_no, u.name, u.line,
               pl.part_no, pl.log_type, pl.qty, pl.updated_at
        FROM part_logs pl
        JOIN users u ON u.employee_no = pl.employee_no
        WHERE pl.log_date = ?
          AND pl.log_time BETWEEN ? AND ?
    """
    params = [selected_date, start_str, end_str]

    if selected_emp:
        query += " AND pl.employee_no = ?"
        params.append(selected_emp)

    if selected_line:
        query += " AND UPPER(u.line) = ?"
        params.append(selected_line)

    query += " ORDER BY u.line, pl.employee_no, pl.log_time, pl.id"

    with db() as conn:
        cur = conn.cursor()
        cur.execute(query, params)
        return cur.fetchall()


def get_summary_by_emp(selected_date: str, start_str: str, end_str: str):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT pl.employee_no, u.name, u.line, SUM(pl.qty) AS total_qty
            FROM part_logs pl
            JOIN users u ON u.employee_no = pl.employee_no
            WHERE pl.log_date = ?
              AND pl.log_time BETWEEN ? AND ?
              AND UPPER(COALESCE(pl.log_type,'GOOD')) IN ('GOOD','FIX')
            GROUP BY pl.employee_no, u.name, u.line
            ORDER BY total_qty DESC
        """, (selected_date, start_str, end_str))
        return cur.fetchall()


def get_summary_by_line(selected_date: str, start_str: str, end_str: str):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT UPPER(u.line) AS line, SUM(pl.qty) AS total_qty
            FROM part_logs pl
            JOIN users u ON u.employee_no = pl.employee_no
            WHERE pl.log_date = ?
              AND pl.log_time BETWEEN ? AND ?
              AND UPPER(COALESCE(pl.log_type,'GOOD')) IN ('GOOD','FIX')
            GROUP BY UPPER(u.line)
            ORDER BY total_qty DESC
        """, (selected_date, start_str, end_str))
        return cur.fetchall()


def get_tech_by_line(selected_date: str, start_str: str, end_str: str):
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT UPPER(u.line) AS line, u.employee_no, u.name,
                   COALESCE(SUM(pl.qty),0) AS total_qty
            FROM users u
            LEFT JOIN part_logs pl ON pl.employee_no = u.employee_no
                 AND pl.log_date = ?
                 AND pl.log_time BETWEEN ? AND ?
                 AND UPPER(COALESCE(pl.log_type,'GOOD')) IN ('GOOD','FIX')
            WHERE u.role='technician' AND u.is_active=1
            GROUP BY UPPER(u.line), u.employee_no, u.name
            ORDER BY UPPER(u.line), total_qty DESC
        """, (selected_date, start_str, end_str))
        return cur.fetchall()


# ==========================================================
# INVENTARIO VIEW POR LINEA (SUPERVISOR)
# ==========================================================
def get_inventory_line_view(selected_date: str, start_str: str, end_str: str):
    with db() as conn:
        cur = conn.cursor()

        cur.execute("""
            SELECT u.employee_no, u.name, UPPER(u.line) AS line,
                   ti.initial_qty
            FROM users u
            LEFT JOIN tech_inventory ti
              ON ti.employee_no = u.employee_no AND ti.log_date = ?
            WHERE u.role='technician' AND u.is_active=1
            ORDER BY UPPER(u.line), u.employee_no
        """, (selected_date,))
        tech_inv_rows = cur.fetchall()

        cur.execute("""
            SELECT employee_no, COALESCE(SUM(qty),0) AS produced
            FROM part_logs
            WHERE log_date = ?
              AND log_time BETWEEN ? AND ?
            GROUP BY employee_no
        """, (selected_date, start_str, end_str))
        produced_rows = cur.fetchall()

    produced_map = {r["employee_no"]: int(r["produced"] or 0) for r in produced_rows}

    inv_by_line: Dict[str, List[dict]] = {}
    inv_lines_map: Dict[str, dict] = {}

    for r in tech_inv_rows:
        emp_no = r["employee_no"]
        name = r["name"]
        line = (r["line"] or "SIN ASIGNAR").upper()

        initial = r["initial_qty"]
        initial = int(initial) if initial is not None else None

        produced = int(produced_map.get(emp_no, 0))  # TODO descuenta inventario
        final = None if initial is None else (initial - produced)

        inv_by_line.setdefault(line, []).append({
            "employee_no": emp_no,
            "name": name,
            "initial": initial,
            "produced": produced,
            "final": final
        })

        if line not in inv_lines_map:
            inv_lines_map[line] = {
                "line": line,
                "initial_sum": 0,
                "produced_sum": 0,
                "final_sum": 0,
                "missing_count": 0,
                "has_any_initial": False
            }

        inv_lines_map[line]["produced_sum"] += produced

        if initial is None:
            inv_lines_map[line]["missing_count"] += 1
        else:
            inv_lines_map[line]["initial_sum"] += initial
            inv_lines_map[line]["final_sum"] += final
            inv_lines_map[line]["has_any_initial"] = True

    inv_lines = []
    for _, data in inv_lines_map.items():
        if not data["has_any_initial"]:
            data["initial_sum"] = None
            data["final_sum"] = None
        del data["has_any_initial"]
        inv_lines.append(data)

    # Garantiza que estén todas las líneas del catálogo
    for ln in [l.upper() for l in LINES]:
        inv_by_line.setdefault(ln, [])
        if not any(x["line"] == ln for x in inv_lines):
            inv_lines.append({
                "line": ln,
                "initial_sum": None,
                "produced_sum": 0,
                "final_sum": None,
                "missing_count": 0
            })

    inv_lines.sort(key=lambda x: x["produced_sum"], reverse=True)

    return inv_lines, inv_by_line


# ==========================================================
# WEEK HELPERS + WEEK QUERIES  (Lunes a Viernes)
# ==========================================================
def parse_year_week_request():
    today = datetime.now(TZ).date()
    default_year = today.isocalendar().year
    default_week = today.isocalendar().week

    year_raw = (request.args.get("year") or str(default_year)).strip()
    week_raw = (request.args.get("week") or str(default_week)).strip()

    year = int(year_raw) if year_raw.isdigit() else default_year
    week = int(week_raw) if week_raw.isdigit() else default_week

    week = max(1, min(53, week))

    try:
        week_start = datetime.fromisocalendar(year, week, 1).date()  # Lunes
        week_end = week_start + timedelta(days=4)  # Viernes
    except Exception:
        year = default_year
        week = default_week
        week_start = datetime.fromisocalendar(year, week, 1).date()
        week_end = week_start + timedelta(days=4)  # Viernes

    return year, week, week_start, week_end


def get_week_totals_emp_map(week_start: date, week_end: date, start_str: str, end_str: str) -> Dict[str, int]:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT employee_no, COALESCE(SUM(qty),0) AS total_qty
            FROM part_logs
            WHERE log_date BETWEEN ? AND ?
              AND log_time BETWEEN ? AND ?
              AND UPPER(COALESCE(log_type,'GOOD')) IN ('GOOD','FIX')
            GROUP BY employee_no
        """, (week_start.isoformat(), week_end.isoformat(), start_str, end_str))
        rows = cur.fetchall()
        return {r["employee_no"]: int(r["total_qty"] or 0) for r in rows}


def get_week_totals_line_map(week_start: date, week_end: date, start_str: str, end_str: str) -> Dict[str, int]:
    with db() as conn:
        cur = conn.cursor()
        cur.execute("""
            SELECT UPPER(u.line) AS line, COALESCE(SUM(pl.qty),0) AS total_qty
            FROM part_logs pl
            JOIN users u ON u.employee_no = pl.employee_no
            WHERE pl.log_date BETWEEN ? AND ?
              AND pl.log_time BETWEEN ? AND ?
              AND UPPER(COALESCE(pl.log_type,'GOOD')) IN ('GOOD','FIX')
            GROUP BY UPPER(u.line)
        """, (week_start.isoformat(), week_end.isoformat(), start_str, end_str))
        rows = cur.fetchall()
        return {(r["line"] or "SIN ASIGNAR").upper(): int(r["total_qty"] or 0) for r in rows}


# ==========================================================
# EXPORT DAILY ROWS
# ==========================================================
def get_export_daily_rows(selected_date: str, start_str: str, end_str: str,
                          selected_emp: str, selected_line: str):
    query = """
        SELECT pl.log_date, pl.log_time, pl.employee_no, u.name, u.line,
               pl.part_no, pl.qty, pl.log_type, pl.updated_at
        FROM part_logs pl
        JOIN users u ON u.employee_no = pl.employee_no
        WHERE pl.log_date = ?
          AND pl.log_time BETWEEN ? AND ?
    """
    params = [selected_date, start_str, end_str]

    if selected_emp:
        query += " AND pl.employee_no = ?"
        params.append(selected_emp)

    if selected_line:
        query += " AND UPPER(u.line) = ?"
        params.append(selected_line)

    query += " ORDER BY u.line, pl.employee_no, pl.log_time, pl.id"

    with db() as conn:
        cur = conn.cursor()
        cur.execute(query, params)
        return cur.fetchall()


# ==========================================================
# ADMIN LOGS LIST
# ==========================================================
def get_admin_logs(selected_date: str, selected_emp: str, selected_line: str):
    query = """
        SELECT pl.id, pl.log_date, pl.log_time, pl.employee_no, u.name, u.line,
               pl.part_no, pl.log_type, pl.qty, pl.updated_at
        FROM part_logs pl
        JOIN users u ON u.employee_no = pl.employee_no
        WHERE pl.log_date = ?
    """
    params = [selected_date]

    if selected_emp:
        query += " AND pl.employee_no = ?"
        params.append(selected_emp)

    if selected_line:
        query += " AND UPPER(u.line) = ?"
        params.append(selected_line)

    query += " ORDER BY pl.log_time ASC, pl.id ASC"

    with db() as conn:
        cur = conn.cursor()
        cur.execute(query, params)
        return cur.fetchall()


# ==========================================================
# ADMIN USER DELETE HELPERS
# ==========================================================
def safe_delete_user(cur: sqlite3.Cursor, employee_no: str) -> dict:
    employee_no = (employee_no or "").strip()

    if not employee_no:
        return {"cat": "err", "msg": "Empleado inválido."}

    if employee_no in {"0000", "9999"}:
        return {"cat": "err", "msg": f"No se puede eliminar el usuario protegido ({employee_no})."}

    if session.get("employee_no") == employee_no:
        return {"cat": "err", "msg": "No puedes eliminar tu propio usuario mientras estás logueado."}

    cur.execute("SELECT employee_no FROM users WHERE employee_no = ?", (employee_no,))
    urow = cur.fetchone()
    if not urow:
        return {"cat": "err", "msg": "Ese usuario no existe."}

    cur.execute("SELECT COUNT(*) AS c FROM part_logs WHERE employee_no = ?", (employee_no,))
    c_logs = int(cur.fetchone()["c"] or 0)

    cur.execute("SELECT COUNT(*) AS c FROM day_closures WHERE employee_no = ?", (employee_no,))
    c_clos = int(cur.fetchone()["c"] or 0)

    cur.execute("SELECT COUNT(*) AS c FROM tech_goals WHERE employee_no = ?", (employee_no,))
    c_goals = int(cur.fetchone()["c"] or 0)

    total_refs = c_logs + c_clos + c_goals

    if total_refs > 0:
        cur.execute("UPDATE users SET is_active = 0 WHERE employee_no = ?", (employee_no,))
        return {
            "cat": "ok",
            "msg": f"El usuario {employee_no} tiene historial (logs/cierres/metas). Se desactivó en lugar de eliminarlo."
        }

    cur.execute("DELETE FROM tech_goals WHERE employee_no = ?", (employee_no,))
    cur.execute("DELETE FROM day_closures WHERE employee_no = ?", (employee_no,))
    cur.execute("DELETE FROM users WHERE employee_no = ?", (employee_no,))
    return {"cat": "ok", "msg": f"Usuario {employee_no} eliminado correctamente."}


def full_delete_user(cur: sqlite3.Cursor, employee_no: str) -> dict:
    employee_no = (employee_no or "").strip()

    if not employee_no:
        return {"cat": "err", "msg": "Empleado inválido."}

    if employee_no in {"0000", "9999"}:
        return {"cat": "err", "msg": f"No se puede eliminar el usuario protegido ({employee_no})."}

    if session.get("employee_no") == employee_no:
        return {"cat": "err", "msg": "No puedes eliminar tu propio usuario mientras estás logueado."}

    cur.execute("SELECT employee_no, role FROM users WHERE employee_no = ?", (employee_no,))
    urow = cur.fetchone()
    if not urow:
        return {"cat": "err", "msg": "Ese usuario no existe."}

    if urow["role"] == "admin":
        return {"cat": "err", "msg": "Por seguridad no se permite eliminar usuarios ADMIN desde aquí."}

    cur.execute("DELETE FROM part_logs WHERE employee_no = ?", (employee_no,))
    cur.execute("DELETE FROM day_closures WHERE employee_no = ?", (employee_no,))
    cur.execute("DELETE FROM tech_goals WHERE employee_no = ?", (employee_no,))
    cur.execute("DELETE FROM users WHERE employee_no = ?", (employee_no,))
    return {"cat": "ok", "msg": f"Usuario {employee_no} eliminado COMPLETAMENTE (incluye historial)."}


# ==========================================================
# EXCEL HELPERS
# ==========================================================
def autosize_worksheets(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            ws.column_dimensions[col_letter].width = min(40, max_len + 2)


# ==========================================================
# MAIN
# ==========================================================
app = create_app()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=APP_PORT, debug=True)

